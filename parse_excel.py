from openpyxl import load_workbook, Workbook
from openpyxl.utils import absolute_coordinate

file_path = ''

def parse_summary_date(file_path):
    # wb = load_workbook(r'C:\Users\test.xlsx')
    # file_path = input('请输入文件路径：')

    wb = load_workbook(file_path)

    global ws
    ws = wb.active

    global head
    head = []

    for row in ws:
        # get the current row number
        row_number = row[0].row
        if row_number != 1:
            continue
        else:
            row_text = ''
            for cell in row:
                row_text += f'{cell.value} '
            head = [cell.value for cell in row]
        # print(row_text)
#     return head, ws

# def separate_data(head, ws):
    '''
        separate each single row data into a new file
    '''
    # thead = ['工号', '姓名', '工资']
    thead = head

    n_wb = Workbook()
    n_ws = n_wb.active

    for i in range(len(thead)):
        n_ws.cell(row=1, column=i+1, value=thead[i])

    for i, row in enumerate(ws):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            # row = 2 because the first row is the head and we only need to generate 1 line
            n_ws.cell(row=2, column=j+1, value=cell.value)
        fileName = row[1].value+'.xlsx'
        n_wb.save(f'D:\{fileName}.xlsx')
        print(f'{fileName} 已生成')


if __name__ == '__main__':
    parse_summary_date(file_path)
    # separate_data(head, ws)

